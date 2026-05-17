"""
URL Validator & Auto-Corrector
==============================
Upload an Excel or CSV file with company names and URLs.
The app validates each URL, attempts auto-correction, and produces a
cleaned file with six new url_ columns ready for enrichment.

Search order for dead/missing URLs:
  1. URL as-is / https:// / www. variants  (no API)
  2. DuckDuckGo instant answer API          (free)
  3. Claude with web_search tool            (Anthropic API, last resort)

Processing model: one row per Streamlit rerun so the Stop button works.
"""

import io
import json
import re
import subprocess
import sys
import time
from difflib import SequenceMatcher
from pathlib import Path
from urllib.parse import urlparse, urlunparse

import anthropic
import pandas as pd
import requests
import streamlit as st

# ── Playwright availability ───────────────────────────────────────────────────
try:
    from playwright.sync_api import sync_playwright
    _PLAYWRIGHT_OK = True
except ImportError:
    _PLAYWRIGHT_OK = False


def _ensure_playwright():
    """Download the Chromium browser binary if not already present."""
    try:
        subprocess.run(
            [sys.executable, "-m", "playwright", "install", "chromium", "--with-deps"],
            capture_output=True,
            timeout=120,
        )
    except Exception:
        pass

# ─────────────────────────────────────────────────────────────────────────────
# Constants
# ─────────────────────────────────────────────────────────────────────────────

_COMPANY_HINTS = ["company", "account", "organisation", "organization", "name", "naam", "bedrijf"]
_DOMAIN_HINTS  = ["domain", "website", "url", "web", "site", "domein"]

_UA      = "Mozilla/5.0 (compatible; URLValidatorBot/1.0)"
_TIMEOUT = 10

MODEL_ID         = "claude-haiku-4-5-20251001"
WEB_SEARCH_TOOL  = {"type": "web_search_20250305", "name": "web_search"}

_COST_INPUT_PER_M  = 0.80
_COST_OUTPUT_PER_M = 4.00

URL_FIELDS = [
    "url_status",
    "final_url",
    "redirect_target",
    "http_status_code",
    "correction_source",
    "correction_notes",
]

_STATUS_COLOR = {
    "ok":          "#1a7a3c",
    "ok_headless": "#0d5c6e",  # teal — confirmed via headless Chrome
    "redirected":  "#b35c00",
    "corrected":   "#7a5c00",
    "unverified":  "#4a3a7a",
    "dead":        "#8b1a1a",
}

_STATUS_TEXT_COLOR = {
    "ok":          "#e6ffe6",
    "ok_headless": "#d6f5ff",
    "redirected":  "#ffe8cc",
    "corrected":   "#fff4cc",
    "unverified":  "#ece6ff",
    "dead":        "#ffe6e6",
}

# ─────────────────────────────────────────────────────────────────────────────
# Utility helpers
# ─────────────────────────────────────────────────────────────────────────────

def _similarity(a: str, b: str) -> float:
    if not a or not b:
        return 0.0
    return SequenceMatcher(None, a.lower(), b.lower()).ratio()


def detect_columns(df: pd.DataFrame):
    cols = df.columns.tolist()
    col_lower = [str(c).lower() for c in cols]

    def best(hints):
        scores = [(max(_similarity(cl, h) for h in hints), i)
                  for i, cl in enumerate(col_lower)]
        score, idx = max(scores)
        return cols[idx], score

    name_col, ns = best(_COMPANY_HINTS)
    domain_col, ds = best(_DOMAIN_HINTS)
    return (
        name_col   if ns >= 0.45 else None,
        domain_col if ds >= 0.55 and domain_col != name_col else None,
    )


def _calc_cost(in_tok: int, out_tok: int) -> float:
    return (in_tok * _COST_INPUT_PER_M + out_tok * _COST_OUTPUT_PER_M) / 1_000_000


def _normalise_url(raw: str) -> str:
    raw = raw.strip()
    if not raw:
        return ""
    if not re.match(r"^https?://", raw, re.IGNORECASE):
        return "https://" + raw
    return raw


def _domain(url: str) -> str:
    try:
        return urlparse(url).netloc.lower().lstrip("www.")
    except Exception:
        return ""


def _get(url: str):
    """GET url; return (response, error_str). Follows redirects."""
    try:
        resp = requests.get(
            url,
            timeout=_TIMEOUT,
            headers={"User-Agent": _UA},
            allow_redirects=True,
        )
        return resp, None
    except requests.exceptions.SSLError:
        try:
            resp = requests.get(
                url,
                timeout=_TIMEOUT,
                headers={"User-Agent": _UA},
                allow_redirects=True,
                verify=False,
            )
            return resp, None
        except Exception as exc:
            return None, str(exc)
    except Exception as exc:
        return None, str(exc)


def check_url_headless(url: str) -> dict:
    """
    Load url in a stealth headless Chromium browser.
    Returns dict: success, status_code, final_url, page_title.
    Never raises — all exceptions become success=False.
    """
    result = {"success": False, "status_code": None, "final_url": url, "page_title": ""}
    if not _PLAYWRIGHT_OK:
        return result
    try:
        with sync_playwright() as pw:
            browser = pw.chromium.launch(headless=True)
            ctx = browser.new_context(
                user_agent=(
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/120.0.0.0 Safari/537.36"
                ),
                viewport={"width": 1280, "height": 800},
                locale="en-US",
                java_script_enabled=True,
                ignore_https_errors=True,
            )
            page = ctx.new_page()
            resp = page.goto(url, timeout=15000, wait_until="domcontentloaded")
            status = resp.status if resp else None
            final  = page.url
            title  = page.title()
            browser.close()

        if status is not None and status < 400:
            result.update(success=True, status_code=status, final_url=final, page_title=title)
        else:
            result.update(status_code=status, final_url=final, page_title=title)
    except Exception:
        pass
    return result


def _get_or_headless(url: str, use_headless: bool):
    """
    Try _get(); if 403/503 and use_headless, retry with headless Chrome.
    Returns (resp, headless_dict_or_None).
    headless_dict is non-None only when headless succeeded.
    """
    resp, _ = _get(url)
    if (
        resp is not None
        and resp.status_code in (403, 503)
        and use_headless
        and _PLAYWRIGHT_OK
    ):
        headless = check_url_headless(url)
        if headless["success"]:
            return resp, headless
    return resp, None


def _parse_json(text: str) -> dict:
    """Extract a JSON object from Claude's response text."""
    text = text.strip()
    # Strip markdown fences if present
    text = re.sub(r"^```(?:json)?\s*", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\s*```$", "", text)
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        # Try to extract the first {...} block
        m = re.search(r"\{.*\}", text, re.DOTALL)
        if m:
            try:
                return json.loads(m.group())
            except json.JSONDecodeError:
                pass
    return {}

# ─────────────────────────────────────────────────────────────────────────────
# Search steps
# ─────────────────────────────────────────────────────────────────────────────

def _duckduckgo_url(company_name: str) -> str | None:
    """Ask DuckDuckGo instant-answer API for the official website."""
    try:
        query = requests.utils.quote(company_name)
        resp = requests.get(
            f"https://api.duckduckgo.com/?q={query}+official+website&format=json&no_redirect=1",
            timeout=_TIMEOUT,
            headers={"User-Agent": _UA},
        )
        if resp.status_code != 200:
            return None
        data = resp.json()
        url = data.get("AbstractURL") or data.get("Redirect") or ""
        if url:
            return url
        for topic in data.get("RelatedTopics", []):
            u = topic.get("FirstURL", "")
            if u and "duckduckgo.com" not in u:
                return u
        return None
    except Exception:
        return None


def find_url_via_claude(company_name: str, api_key: str) -> dict:
    """
    Use Claude + web_search_20250305 to find the official website URL.

    Returns:
        url      – full working URL or empty string
        notes    – one-sentence explanation
        verified – whether we confirmed the URL with a GET request
        http_status – HTTP status of the verified URL (or "")
        tokens_in, tokens_out – usage for cost tracking
    """
    result = {
        "url": "",
        "notes": "",
        "verified": False,
        "http_status": "",
        "tokens_in": 0,
        "tokens_out": 0,
    }

    prompt = (
        f"Search for the official website of '{company_name}'.\n"
        "You MUST return a JSON object with exactly these two fields:\n"
        "- url: the complete working URL you found, always starting with https:// "
        "(example: https://www.abnamro.nl). This field must NEVER be empty if you found a website.\n"
        "- notes: one short sentence describing what you found.\n"
        "If and only if the company truly has no website or is dissolved, "
        "set url to empty string and explain in notes.\n"
        "Return ONLY raw JSON. No markdown, no explanation outside the JSON."
    )

    try:
        client   = anthropic.Anthropic(api_key=api_key)
        messages = [{"role": "user", "content": prompt}]
        total_in = total_out = 0

        for _ in range(6):  # cap agentic loop iterations
            resp = client.messages.create(
                model=MODEL_ID,
                max_tokens=256,
                tools=[WEB_SEARCH_TOOL],
                messages=messages,
            )
            total_in  += resp.usage.input_tokens
            total_out += resp.usage.output_tokens

            if resp.stop_reason == "end_turn":
                text = "".join(getattr(b, "text", "") for b in resp.content).strip()
                break

            if resp.stop_reason == "tool_use":
                messages.append({"role": "assistant", "content": resp.content})
                tool_results = [
                    {"type": "tool_result", "tool_use_id": b.id, "content": ""}
                    for b in resp.content
                    if getattr(b, "type", "") == "tool_use"
                ]
                if tool_results:
                    messages.append({"role": "user", "content": tool_results})
            else:
                text = "".join(getattr(b, "text", "") for b in resp.content).strip()
                break
        else:
            text = ""

        result["tokens_in"]  = total_in
        result["tokens_out"] = total_out

        parsed = _parse_json(text)
        url    = (parsed.get("url") or "").strip()
        notes  = (parsed.get("notes") or "").strip()
        result["notes"] = notes or "Claude returned no explanation"

        # Fallback: if url field is empty, try to extract one from the notes text
        if not url and notes:
            m = re.search(r'https?://[^\s\'"<>]+', notes)
            if m:
                url = m.group().rstrip(".,)")

        if url:
            url = _normalise_url(url)
            verify_resp, _ = _get(url)
            if verify_resp is not None and verify_resp.status_code < 400:
                result["url"]         = verify_resp.url
                result["verified"]    = True
                result["http_status"] = verify_resp.status_code
            else:
                # URL not verified (bot-blocking etc.) — save it anyway
                result["url"]         = url
                result["verified"]    = False
                result["http_status"] = getattr(verify_resp, "status_code", "")

    except anthropic.APIError as exc:
        result["notes"] = f"Claude API error: {exc}"

    return result

# ─────────────────────────────────────────────────────────────────────────────
# Main validation waterfall
# ─────────────────────────────────────────────────────────────────────────────

def validate_url(
    original_url: str,
    company_name: str,
    api_key: str = "",
    use_claude: bool = False,
    use_headless: bool = False,
) -> dict:
    """
    Validate one URL through the waterfall.
    Returns the six url_ fields plus tokens_in / tokens_out.
    """
    result = {
        "url_status":       "dead",
        "final_url":        original_url,
        "redirect_target":  "",
        "http_status_code": "",
        "correction_source":"dead",
        "correction_notes": "",
        "tokens_in":        0,
        "tokens_out":       0,
    }

    raw = (original_url or "").strip()
    if not raw:
        result["correction_notes"] = "no URL provided"
        return result

    # ── Step 1a: URL as-is ────────────────────────────────────────────────────
    url_as_is = _normalise_url(raw)
    resp, headless = _get_or_headless(url_as_is, use_headless)
    if headless:
        result.update(
            url_status="ok_headless",
            final_url=headless["final_url"],
            http_status_code=headless["status_code"],
            correction_source="headless_chrome",
            correction_notes=f"bot-blocked ({resp.status_code}), confirmed via headless Chrome",
        )
        return result
    if resp is not None and resp.status_code < 400:
        final     = resp.url
        orig_dom  = _domain(url_as_is)
        final_dom = _domain(final)
        result["http_status_code"] = resp.status_code
        result["final_url"]        = final
        if orig_dom and final_dom and orig_dom != final_dom:
            result["url_status"]        = "redirected"
            result["redirect_target"]   = final
            result["correction_source"] = "original"
            result["correction_notes"]  = f"redirected to {final_dom}"
        else:
            result["url_status"]        = "ok"
            result["correction_source"] = "original"
        return result

    # ── Step 1b: add https:// if missing ─────────────────────────────────────
    if not re.match(r"^https?://", raw, re.IGNORECASE):
        url_https = "https://" + raw
        resp, headless = _get_or_headless(url_https, use_headless)
        if headless:
            result.update(
                url_status="ok_headless",
                final_url=headless["final_url"],
                http_status_code=headless["status_code"],
                correction_source="headless_chrome",
                correction_notes=f"bot-blocked ({resp.status_code}), confirmed via headless Chrome",
            )
            return result
        if resp is not None and resp.status_code < 400:
            final = resp.url
            result["http_status_code"]  = resp.status_code
            result["final_url"]         = final
            result["url_status"]        = "corrected"
            result["correction_source"] = "https_added"
            result["correction_notes"]  = "added https:// prefix"
            if _domain(url_https) != _domain(final):
                result["redirect_target"] = final
            return result

    # ── Step 1c: www. variant ─────────────────────────────────────────────────
    parsed = urlparse(url_as_is)
    if not parsed.netloc.lower().startswith("www."):
        url_www = urlunparse(parsed._replace(netloc="www." + parsed.netloc))
        resp, headless = _get_or_headless(url_www, use_headless)
        if headless:
            result.update(
                url_status="ok_headless",
                final_url=headless["final_url"],
                http_status_code=headless["status_code"],
                correction_source="headless_chrome",
                correction_notes=f"bot-blocked ({resp.status_code}), confirmed via headless Chrome",
            )
            return result
        if resp is not None and resp.status_code < 400:
            final = resp.url
            result["http_status_code"]  = resp.status_code
            result["final_url"]         = final
            result["url_status"]        = "corrected"
            result["correction_source"] = "www_added"
            result["correction_notes"]  = "added www. prefix"
            if _domain(url_www) != _domain(final):
                result["redirect_target"] = final
            return result

    # ── Step 2: DuckDuckGo instant answer (free) ──────────────────────────────
    if company_name:
        ddg_url = _duckduckgo_url(company_name)
        if ddg_url:
            resp, _ = _get(ddg_url)
            if resp is not None and resp.status_code < 400:
                final = resp.url
                result["http_status_code"]  = resp.status_code
                result["final_url"]         = final
                result["url_status"]        = "corrected"
                result["correction_source"] = "search_corrected"
                result["correction_notes"]  = f"found via DuckDuckGo: {_domain(final)}"
                if _domain(ddg_url) != _domain(final):
                    result["redirect_target"] = final
                return result

    # ── Step 3: Claude web search (last resort, costs tokens) ─────────────────
    if use_claude and api_key and company_name:
        claude_res = find_url_via_claude(company_name, api_key)
        result["tokens_in"]  = claude_res["tokens_in"]
        result["tokens_out"] = claude_res["tokens_out"]

        if claude_res["url"]:
            final = claude_res["url"]
            result["http_status_code"]  = claude_res["http_status"]
            result["final_url"]         = final
            result["correction_source"] = "search_corrected"
            result["correction_notes"]  = f"Claude: {claude_res['notes']}"
            if _domain(final) != _domain(url_as_is):
                result["redirect_target"] = final
            if claude_res["verified"]:
                result["url_status"] = "corrected"
            else:
                # Claude found a URL but our GET was blocked — mark unverified,
                # not dead, so the user can decide rather than lose the lead
                result["url_status"] = "unverified"
            return result
        elif claude_res["notes"]:
            result["correction_notes"] = f"Claude: {claude_res['notes']}"

    # ── All steps failed ──────────────────────────────────────────────────────
    result["url_status"]        = "dead"
    result["correction_source"] = "dead"
    if not result["correction_notes"]:
        result["correction_notes"] = "no working URL found"
    return result


# ─────────────────────────────────────────────────────────────────────────────
# Excel builders
# ─────────────────────────────────────────────────────────────────────────────

def _build_cleaned_excel(out_df: pd.DataFrame) -> bytes:
    """Standard cleaned Excel — all original columns + url_ columns."""
    buf = io.BytesIO()
    out_df.to_excel(buf, index=False, engine="openpyxl")
    return buf.getvalue()


def _build_manual_excel(out_df: pd.DataFrame, url_col: str) -> bytes:
    """
    Manual-editing Excel:
    - dead rows: Website URL column cleared + yellow highlight + comment
    - unverified rows: final_url used + light-orange highlight
    - ok/corrected/redirected: final_url used
    - header row bolded, columns auto-fitted, A1 note added
    """
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.comments import Comment

    YELLOW      = PatternFill("solid", fgColor="FFFF00")
    ORANGE      = PatternFill("solid", fgColor="FFE0B2")
    HEADER_FONT = Font(bold=True)

    df = out_df.copy()

    # Resolve Website URL column: use final_url for non-dead rows
    if "final_url" in df.columns and url_col in df.columns:
        mask_good = df["url_status"].isin(["ok", "corrected", "redirected", "unverified"])
        df.loc[mask_good, url_col] = df.loc[mask_good, "final_url"]
        df.loc[df["url_status"] == "dead", url_col] = ""

    wb = Workbook()
    ws = wb.active

    cols = df.columns.tolist()
    ws.append(cols)

    # Bold + centre header
    for cell in ws[1]:
        cell.font      = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    # A1 legend comment
    legend = Comment(
        "Yellow = URL missing, fill manually. Orange = URL unverified (may be bot-blocked).",
        "URL Validator",
    )
    legend.width  = 320
    legend.height = 60
    ws["A1"].comment = legend

    url_col_idx = cols.index(url_col) + 1 if url_col in cols else None  # 1-based

    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        status = row.get("url_status", "")
        ws.append([row[c] for c in cols])

        if status == "dead":
            for cell in ws[row_idx]:
                cell.fill = YELLOW
            if url_col_idx:
                url_cell = ws.cell(row=row_idx, column=url_col_idx)
                url_cell.comment = Comment(
                    "URL not found — please fill in manually",
                    "URL Validator",
                )
        elif status == "unverified":
            for cell in ws[row_idx]:
                cell.fill = ORANGE

    # Auto-fit column widths (cap at 60)
    for col_cells in ws.columns:
        max_len = max(
            (len(str(c.value)) if c.value is not None else 0) for c in col_cells
        )
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 60)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# Streamlit UI
# ─────────────────────────────────────────────────────────────────────────────

st.set_page_config(page_title="URL Validator", page_icon="🔗", layout="wide")
st.title("🔗 URL Validator & Auto-Corrector")
st.caption("Validate and fix company URLs before enrichment.")

# ── API key ───────────────────────────────────────────────────────────────────
_api_key = ""
try:
    _api_key = st.secrets.get("ANTHROPIC_API_KEY", "") or ""
except Exception:
    pass

# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.header("Settings")

    # API key status
    if _api_key:
        st.success("API key loaded", icon="✅")
    else:
        st.warning("ANTHROPIC_API_KEY not set in secrets", icon="⚠️")

    use_claude = st.toggle(
        "Use Claude web search for dead URLs",
        value=bool(_api_key),
        disabled=not bool(_api_key),
        help="Uses claude-haiku-4-5 with web_search as a last resort. Costs ~$0.001 per lookup.",
    )

    st.divider()

    if _PLAYWRIGHT_OK:
        use_headless = st.toggle(
            "Use headless Chrome for bot-blocked URLs",
            value=True,
            help="Retries 403/503 responses with a stealth Chromium browser.",
        )
        st.caption("⚠ First run may take 60s to install Chromium")
    else:
        use_headless = False
        st.caption("Headless Chrome unavailable (playwright not installed)")

    st.divider()
    st.subheader("Token usage")
    tok_in_ph  = st.empty()
    tok_out_ph = st.empty()
    cost_ph    = st.empty()

# ── Session-state initialisation ─────────────────────────────────────────────
for key, default in {
    "_running":           False,
    "_stop":              False,
    "_current_idx":       0,
    "_results":           [],
    "_df_raw":            None,
    "_name_col":          None,
    "_url_col":           None,
    "_tokens_in":         0,
    "_tokens_out":        0,
    "_playwright_ready":  False,
}.items():
    if key not in st.session_state:
        st.session_state[key] = default

# Ensure Chromium binary is present (once per session)
if _PLAYWRIGHT_OK and not st.session_state["_playwright_ready"]:
    _ensure_playwright()
    st.session_state["_playwright_ready"] = True


def _reset():
    st.session_state["_running"]     = False
    st.session_state["_stop"]        = False
    st.session_state["_current_idx"] = 0
    st.session_state["_results"]     = []
    st.session_state["_tokens_in"]   = 0
    st.session_state["_tokens_out"]  = 0
    # keep _playwright_ready — no need to reinstall Chromium on data reset


def _render_token_sidebar():
    ti = st.session_state["_tokens_in"]
    to = st.session_state["_tokens_out"]
    cost = _calc_cost(ti, to)
    tok_in_ph.metric("Input tokens",  f"{ti:,}")
    tok_out_ph.metric("Output tokens", f"{to:,}")
    cost_ph.metric("Estimated cost",  f"${cost:.4f}")


_render_token_sidebar()

# ── File upload ───────────────────────────────────────────────────────────────
uploaded = st.file_uploader("Upload Excel or CSV file", type=["xlsx", "xls", "csv"])

if uploaded:
    try:
        if uploaded.name.lower().endswith(".csv"):
            df_raw = pd.read_csv(uploaded)
        else:
            df_raw = pd.read_excel(uploaded)
        df_raw = df_raw.reset_index(drop=True)
    except Exception as exc:
        st.error(f"Could not read file: {exc}")
        st.stop()

    if st.session_state["_df_raw"] is None or not df_raw.equals(st.session_state["_df_raw"]):
        _reset()
        st.session_state["_df_raw"] = df_raw

    df_raw = st.session_state["_df_raw"]
    auto_name, auto_url = detect_columns(df_raw)

    st.subheader("Column selection")
    col_a, col_b = st.columns(2)
    with col_a:
        name_col = st.selectbox(
            "Company name column",
            options=df_raw.columns.tolist(),
            index=df_raw.columns.tolist().index(auto_name) if auto_name else 0,
        )
    with col_b:
        url_col = st.selectbox(
            "URL column",
            options=df_raw.columns.tolist(),
            index=df_raw.columns.tolist().index(auto_url) if auto_url else 0,
        )

    st.session_state["_name_col"] = name_col
    st.session_state["_url_col"]  = url_col

    st.subheader("Data preview")
    st.dataframe(df_raw.head(10), use_container_width=True)

    # ── Controls ──────────────────────────────────────────────────────────────
    total = len(df_raw)
    ctrl_l, ctrl_r = st.columns([1, 1])
    with ctrl_l:
        start_btn = st.button(
            "▶ Start validation",
            disabled=st.session_state["_running"],
            type="primary",
        )
    with ctrl_r:
        stop_btn = st.button(
            "⏹ Stop",
            disabled=not st.session_state["_running"],
        )

    if start_btn:
        _reset()
        st.session_state["_running"] = True
        st.rerun()

    if stop_btn:
        st.session_state["_stop"] = True

    # ── Processing loop (one row per rerun) ───────────────────────────────────
    if st.session_state["_running"]:
        idx = st.session_state["_current_idx"]

        if st.session_state["_stop"] or idx >= total:
            st.session_state["_running"] = False
            if st.session_state["_stop"]:
                st.info(f"Stopped after {idx} row(s).")
            st.rerun()

        st.progress(idx / total, text=f"Processing row {idx + 1} of {total}…")

        row     = df_raw.iloc[idx]
        company = str(row.get(name_col, "") or "")
        url     = str(row.get(url_col,  "") or "")

        url_result = validate_url(
            url, company,
            api_key=_api_key,
            use_claude=use_claude,
            use_headless=use_headless,
        )

        # Accumulate token usage
        st.session_state["_tokens_in"]  += url_result.pop("tokens_in",  0)
        st.session_state["_tokens_out"] += url_result.pop("tokens_out", 0)
        _render_token_sidebar()

        st.session_state["_results"].append({**row.to_dict(), **url_result})
        st.session_state["_current_idx"] += 1

        if st.session_state["_results"]:
            live_df = pd.DataFrame(st.session_state["_results"])
            st.subheader(f"Live results ({len(live_df)} processed)")
            _display_cols = [url_col, "url_status", "final_url", "http_status_code",
                             "correction_source", "correction_notes"]
            _display_cols = [c for c in _display_cols if c in live_df.columns]

            def _row_style(row):
                status = row.get("url_status", "")
                bg = _STATUS_COLOR.get(status, "")
                fg = _STATUS_TEXT_COLOR.get(status, "")
                style = f"background-color: {bg}; color: {fg}" if bg else ""
                return [style for _ in row]

            st.dataframe(
                live_df[_display_cols].style.apply(_row_style, axis=1),
                use_container_width=True,
            )

        time.sleep(0.05)
        st.rerun()

    # ── Results panel ─────────────────────────────────────────────────────────
    results = st.session_state["_results"]
    if results and not st.session_state["_running"]:
        results_df = pd.DataFrame(results)

        st.subheader("Summary")
        counts = results_df["url_status"].value_counts()
        m1, m2, m3, m4, m5, m6 = st.columns(6)
        m1.metric("✅ OK",             counts.get("ok", 0))
        m2.metric("🌐 Headless OK",    counts.get("ok_headless", 0))
        m3.metric("↩️ Redirected",     counts.get("redirected", 0))
        m4.metric("🔧 Corrected",      counts.get("corrected", 0))
        m5.metric("🔮 Unverified",     counts.get("unverified", 0))
        m6.metric("💀 Dead",           counts.get("dead", 0))

        unverified_count = counts.get("unverified", 0)
        if unverified_count:
            st.info(
                f"🔮 {unverified_count} URL(s) were found by Claude but could not be verified "
                "(site may block bots). Check them manually — they are likely correct."
            )
        dead_count = counts.get("dead", 0)
        if dead_count:
            st.warning(
                f"⚠️ {dead_count} URL(s) could not be resolved. "
                "These rows are highlighted in red — fix them manually before running enrichment."
            )

        st.subheader("Full results")
        display_cols = [name_col, url_col] + URL_FIELDS
        display_cols = [c for c in display_cols if c in results_df.columns]

        def _row_style(row):
            status = row.get("url_status", "")
            bg = _STATUS_COLOR.get(status, "")
            fg = _STATUS_TEXT_COLOR.get(status, "")
            style = f"background-color: {bg}; color: {fg}" if bg else ""
            return [style for _ in row]

        st.dataframe(
            results_df[display_cols].style.apply(_row_style, axis=1),
            use_container_width=True,
            height=500,
        )

        dead_df = results_df[results_df["url_status"] == "dead"]
        if not dead_df.empty:
            with st.expander(f"💀 Dead URLs ({len(dead_df)} rows) — needs manual fix"):
                st.dataframe(
                    dead_df[[name_col, url_col, "correction_notes"]],
                    use_container_width=True,
                )

        st.subheader("Download")
        orig_cols = df_raw.columns.tolist()
        new_cols  = [c for c in URL_FIELDS if c not in orig_cols]
        out_df    = (
            results_df[orig_cols + new_cols]
            if all(c in results_df.columns for c in orig_cols)
            else results_df
        )

        # Derive stem from uploaded filename
        stem = Path(uploaded.name).stem

        dl_l, dl_r = st.columns(2)
        with dl_l:
            st.download_button(
                label="⬇️ Download cleaned Excel",
                data=_build_cleaned_excel(out_df),
                file_name="validated_urls.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        with dl_r:
            st.download_button(
                label="✏️ Download for manual editing",
                data=_build_manual_excel(out_df, url_col),
                file_name=f"{stem}_for_manual_editing.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        # Summary line
        n_ready      = (counts.get("ok", 0) + counts.get("ok_headless", 0)
                        + counts.get("corrected", 0) + counts.get("redirected", 0))
        n_dead       = counts.get("dead", 0)
        n_unverified = counts.get("unverified", 0)
        st.caption(
            f"Ready to enrich: **{n_ready}** companies | "
            f"Needs manual URL: **{n_dead}** companies | "
            f"Unverified (likely ok): **{n_unverified}** companies"
        )

else:
    st.info("Upload an Excel or CSV file to get started.")
